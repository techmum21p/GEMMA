import json
import os
import tempfile
from datetime import datetime, date
from pathlib import Path

import pytest

from app.services.export_service import generate_excel_report
from app.services.pdf_service import generate_pdf


MOCK_SHIFT = {
    "id": "test-shift-001",
    "bhw_name": "Maria Santos",
    "date": date(2026, 5, 1),
    "start_time": datetime(2026, 5, 1, 8, 0),
    "end_time": datetime(2026, 5, 1, 16, 0),
    "coordinator_email": "coordinator@health.gov.ph",
}

MOCK_PATIENTS = [
    {
        "id": 1,
        "shift_id": "test-shift-001",
        "timestamp": datetime(2026, 5, 1, 8, 30),
        "name": "Juan dela Cruz",
        "age": 35,
        "sex": "M",
        "chief_complaint": "Masakit ang ulo at may lagnat.",
        "image_path": None,
        "image_findings": None,
        "followup_qa": "{}",
        "triage_level": "YELLOW",
        "top_conditions": json.dumps([
            {"rank": 1, "condition": "Flu", "plain_explanation": "Karaniwang lagnat."},
        ]),
        "handoff_summary": json.dumps({"S": "Lagnat.", "O": "—", "A": "Flu.", "P": "Rest."}),
        "status": "Seen",
        "pdf_path": None,
        "triage_reason": "Di urgent pero kailangan ng konsultasyon.",
        "disclaimer": "Para sa kaalaman ng BHW lamang.",
    },
    {
        "id": 2,
        "shift_id": "test-shift-001",
        "timestamp": datetime(2026, 5, 1, 9, 0),
        "name": None,
        "age": 60,
        "sex": "F",
        "chief_complaint": "Nanghihina at nahihirapan huminga.",
        "image_path": None,
        "image_findings": None,
        "followup_qa": "{}",
        "triage_level": "RED",
        "top_conditions": json.dumps([
            {"rank": 1, "condition": "Heart attack", "plain_explanation": "Posibleng problema sa puso."},
        ]),
        "handoff_summary": json.dumps({"S": "Nahihirapan huminga.", "O": "—", "A": "Posible heart attack.", "P": "I-refer agad."}),
        "status": "Referred",
        "pdf_path": None,
        "triage_reason": "Agarang atensyon kailangan.",
        "disclaimer": "Para sa kaalaman ng BHW lamang.",
    },
]


def test_excel_report_creates_file():
    excel_path = generate_excel_report(MOCK_SHIFT, MOCK_PATIENTS)
    assert Path(excel_path).exists()
    assert excel_path.endswith(".xlsx")
    os.unlink(excel_path)


def test_excel_report_correct_sheets():
    import openpyxl
    excel_path = generate_excel_report(MOCK_SHIFT, MOCK_PATIENTS)
    wb = openpyxl.load_workbook(excel_path)
    assert "Patient Log" in wb.sheetnames
    assert "Shift Summary" in wb.sheetnames
    os.unlink(excel_path)


def test_excel_report_patient_count():
    import openpyxl
    excel_path = generate_excel_report(MOCK_SHIFT, MOCK_PATIENTS)
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["Patient Log"]
    assert ws.max_row == len(MOCK_PATIENTS) + 1  # +1 for header
    os.unlink(excel_path)


def test_pdf_report_creates_file():
    patient = MOCK_PATIENTS[0]
    pdf_path = generate_pdf(patient, "Maria Santos")
    assert Path(pdf_path).exists()
    assert pdf_path.endswith(".pdf")
    os.unlink(pdf_path)
