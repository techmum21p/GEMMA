import json
import logging
from collections import Counter
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

EXPORT_DIR = Path("exports/reports")
EXPORT_DIR.mkdir(parents=True, exist_ok=True)

TRIAGE_COLORS = {
    "RED": "C0392B",
    "YELLOW": "E67E22",
    "GREEN": "2D6A2D",
}


def generate_excel_report(shift: dict, patients: list[dict]) -> str:
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    safe_name = shift.get("bhw_name", "BHW").replace(" ", "_")
    file_path = EXPORT_DIR / f"shift_{safe_name}_{ts}.xlsx"

    rows = []
    for i, p in enumerate(patients, 1):
        top_conditions = p.get("top_conditions", "[]")
        if isinstance(top_conditions, str):
            try:
                conditions = json.loads(top_conditions)
            except Exception:
                conditions = []
        else:
            conditions = top_conditions

        top_condition = conditions[0]["condition"] if conditions else "—"

        timestamp = p.get("timestamp", "")
        if hasattr(timestamp, "strftime"):
            time_str = timestamp.strftime("%H:%M")
        else:
            time_str = str(timestamp)[:16]

        rows.append({
            "#": i,
            "Oras": time_str,
            "Pangalan": p.get("name") or "—",
            "Edad": p.get("age") or "—",
            "Kasarian": p.get("sex") or "—",
            "Chief Complaint": p.get("chief_complaint", ""),
            "Triage Level": p.get("triage_level", ""),
            "Nangungunang Kondisyon": top_condition,
            "Status": p.get("status", "Pending"),
        })

    df_patients = pd.DataFrame(rows)

    total = len(patients)
    red = sum(1 for p in patients if p.get("triage_level") == "RED")
    yellow = sum(1 for p in patients if p.get("triage_level") == "YELLOW")
    green = sum(1 for p in patients if p.get("triage_level") == "GREEN")

    all_conditions = []
    for p in patients:
        top_conditions = p.get("top_conditions", "[]")
        if isinstance(top_conditions, str):
            try:
                conditions = json.loads(top_conditions)
            except Exception:
                conditions = []
        else:
            conditions = top_conditions
        if conditions:
            all_conditions.append(conditions[0]["condition"])

    top3 = Counter(all_conditions).most_common(3)

    start_time = shift.get("start_time", "")
    end_time = shift.get("end_time", "")
    if hasattr(start_time, "strftime"):
        start_str = start_time.strftime("%Y-%m-%d %H:%M")
    else:
        start_str = str(start_time)[:16]
    if end_time and hasattr(end_time, "strftime"):
        end_str = end_time.strftime("%Y-%m-%d %H:%M")
    elif end_time:
        end_str = str(end_time)[:16]
    else:
        end_str = "Hindi pa tapos"

    summary_rows = [
        ["Impormasyon ng Shift", ""],
        ["BHW", shift.get("bhw_name", "")],
        ["Petsa", str(shift.get("date", ""))],
        ["Simula ng Shift", start_str],
        ["Katapusan ng Shift", end_str],
        ["", ""],
        ["Bilang ng Pasyente", ""],
        ["Kabuuan", total],
        ["🔴 RED", f"{red} ({red/total*100:.0f}%)" if total else "0 (0%)"],
        ["🟡 YELLOW", f"{yellow} ({yellow/total*100:.0f}%)" if total else "0 (0%)"],
        ["🟢 GREEN", f"{green} ({green/total*100:.0f}%)" if total else "0 (0%)"],
        ["", ""],
        ["Top 3 Kondisyon", ""],
    ]
    for rank, (condition, count) in enumerate(top3, 1):
        summary_rows.append([f"#{rank} {condition}", f"{count} pasyente"])

    df_summary = pd.DataFrame(summary_rows, columns=["Kategorya", "Halaga"])

    with pd.ExcelWriter(str(file_path), engine="openpyxl") as writer:
        df_patients.to_excel(writer, sheet_name="Patient Log", index=False)
        df_summary.to_excel(writer, sheet_name="Shift Summary", index=False, header=False)

        _style_patient_sheet(writer.sheets["Patient Log"], df_patients)
        _style_summary_sheet(writer.sheets["Shift Summary"])

    logger.info(f"Excel report generated: {file_path}")
    return str(file_path)


def _style_patient_sheet(ws, df: pd.DataFrame) -> None:
    header_fill = PatternFill(start_color="1B3A6B", end_color="1B3A6B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    for col_num, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    triage_fills = {k: PatternFill(start_color=v, end_color=v, fill_type="solid") for k, v in TRIAGE_COLORS.items()}

    for row_num in range(2, ws.max_row + 1):
        triage_cell = ws.cell(row=row_num, column=7)
        level = str(triage_cell.value or "")
        if level in triage_fills:
            triage_cell.fill = triage_fills[level]
            triage_cell.font = Font(color="FFFFFF", bold=True)

    col_widths = [4, 8, 20, 6, 8, 35, 14, 30, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def _style_summary_sheet(ws) -> None:
    bold_rows = {1, 7, 13}
    header_fill = PatternFill(start_color="1B3A6B", end_color="1B3A6B", fill_type="solid")

    for row_num in range(1, ws.max_row + 1):
        cell_a = ws.cell(row=row_num, column=1)
        if row_num in bold_rows:
            cell_a.fill = header_fill
            cell_a.font = Font(color="FFFFFF", bold=True)
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 25
